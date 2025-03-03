from openpyxl import load_workbook
import pandas as pd
import tempfile
import os
import shutil
from pandas.io.formats import excel

excel.ExcelFormatter.header_style = None
global xl_interface, interface_name

def load_workbook_openpyxl(filepath, read_only=False, data_only=True):
    temp_dir = None
    workbook = None
    try:
        if read_only:
            temp_dir = tempfile.mkdtemp()
            temp_file_path = os.path.join(temp_dir, os.path.basename(filepath))
            shutil.copy(filepath, temp_file_path)
            workbook = load_workbook(temp_file_path, read_only=read_only, data_only=True)
        else:
            workbook = load_workbook(filename=filepath, data_only=data_only, read_only=read_only)
    except Exception as e:
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)

        # Close the workbook explicitly if opened in read-only mode
        if workbook and read_only:
            workbook.close()

        raise Exception(f"Error loading workbook: {e}")

    return workbook

def close_workbook(workbook, temp_dir = None):
    if temp_dir:
        shutil.rmtree(temp_dir, ignore_errors=True)

    # Close the workbook explicitly if opened in read-only mode
    if workbook:
        workbook._archive.close()
        workbook.close()

def load_excel_config(ref_file_path):
    global config_file, config_file_name
    config_file = load_workbook_openpyxl(ref_file_path, read_only=True, data_only=True)
    config_file_name = ref_file_path

def get_dataframe_from_range(workbook, named_range, header=True):
    data = get_named_range(workbook, named_range)
    df = RangeToDataframe(data, header=header)
    return df

def get_named_range(workbook, named_range):
    start_cell_ref = workbook.defined_names[named_range].attr_text
    sheet_name, start_cell_address = start_cell_ref.split('!')

    # Remove quotes from sheet name if present
    sheet_name = sheet_name.strip("'")
    sheet = workbook[sheet_name]

    # Get the row and column of the starting cell
    start_cell = sheet[start_cell_address]
    start_row = start_cell.row
    start_col = start_cell.column

    # Extract all rows from the starting point onwards
    first_row = \
    list(sheet.iter_rows(min_row=start_row, max_row=start_row, min_col=start_col, values_only=True))[0]
    for col_idx, value in enumerate(first_row, start=start_col):
        if value is None:
            last_col = col_idx - 1
            break
    else:
        last_col = start_col + len(first_row) - 1

    # Determine last row by scanning the first column once
    for row_idx, value in enumerate(sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col,
                                                    max_col=start_col, values_only=True), start=start_row):
        if value[0] is None:
            last_row = row_idx - 1
            break
    else:
        last_row = start_row + len(list(sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col,
                                                        max_col=start_col, values_only=True))) - 1

    # Extract data within the defined range
    rows = sheet.iter_rows(min_row=start_row, max_row=last_row, min_col=start_col, max_col=last_col, values_only=True)

    data = [list(row) for row in rows]

    return data


def RangeToDataframe(data, header=False, replace_datetime=False, cols=[]):
    if header:
        df = pd.DataFrame(data[1:], columns=data[0])
    else:
        df = pd.DataFrame(data)
        if cols != []:
            df.columns = cols
    return df

def range_to_dataframe(data):
    # Print the non-empty data
    headers = data[0]
    data_rows = data[1:]
    # Create a Pandas DataFrame
    df = pd.DataFrame(data_rows, columns=headers)
    return df


def get_value_from_named_ranged(workbook, named_range):
    start_cell_ref = workbook.defined_names[named_range].attr_text
    sheet_name, start_cell_address = start_cell_ref.split('!')

    # Remove quotes from sheet name if present
    sheet_name = sheet_name.strip("'")
    sheet = workbook[sheet_name]
    value = sheet[start_cell_address].value
    # Get the row and column of the starting cell
    return value

def write_to_excel(df, excel_file_path, sheet_name ="", target_address = (), named_range_name = "", mode='a',
                   header=True):
    # Load the workbook
    wb = load_workbook(excel_file_path)

    # Locate the named range
    if named_range_name in wb.defined_names and named_range_name != "":
        # Get the range coordinates (e.g., Sheet1!$A$4)
        dest_range = wb.defined_names[named_range_name].attr_text
        sheet_name, cell_reference = dest_range.split("!")
        cell_reference = cell_reference.replace('$', '')  # Remove absolute reference symbols

        # Calculate starting row and column
        startrow = int(''.join(filter(str.isdigit, cell_reference))) - 1  # Extract and convert row number
        startcol = ord(''.join(filter(str.isalpha, cell_reference)).upper()) - 65  # Extract and convert column letter

    else:
        startrow = target_address[0]
        startcol = target_address[1]
        # Use Pandas ExcelWriter with the existing workbook

    with pd.ExcelWriter(excel_file_path,  if_sheet_exists="overlay", engine="openpyxl", mode=mode) as writer:
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, startcol=startcol, index=False, header=header)

